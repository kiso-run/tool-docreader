"""tool-docreader — extract text from PDF, DOCX, CSV, XLSX, and plain text files.

Subprocess contract (same as all kiso tools):
  stdin:  JSON {args, session, workspace, session_secrets, plan_outputs}
  stdout: result text on success
  stderr: error description on failure
  exit 0: success, exit 1: failure
"""
from __future__ import annotations

import csv
import io
import json
import signal
import sys
from pathlib import Path

signal.signal(signal.SIGTERM, lambda *_: sys.exit(0))

# Maximum characters to output.  Each format reader self-truncates at semantic
# boundaries (page, row, paragraph) so the planner gets actionable continuation
# hints instead of a mid-text chop.
_MAX_OUTPUT_CHARS = 50_000

# Plain text extensions (read as-is).
_TEXT_EXTENSIONS = frozenset({
    ".txt", ".md", ".rst", ".log", ".json", ".yaml", ".yml",
    ".toml", ".ini", ".cfg", ".conf", ".sh", ".bash", ".py",
    ".js", ".ts", ".html", ".htm", ".xml", ".css", ".sql",
})


def main() -> None:
    data = json.load(sys.stdin)
    args = data.get("args", {})
    workspace = data.get("workspace", ".")

    action = args.get("action", "read")

    try:
        if action == "list":
            result = do_list(workspace)
        elif action == "info":
            result = do_info(workspace, args)
        elif action == "read":
            result = do_read(workspace, args)
        else:
            print(f"Unknown action: {action}", file=sys.stderr)
            sys.exit(1)
    except FileNotFoundError as e:
        print(f"File not found: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

    print(result)


# ---------------------------------------------------------------------------
# Actions
# ---------------------------------------------------------------------------


def do_list(workspace: str) -> str:
    """List files in the uploads/ directory."""
    uploads = Path(workspace) / "uploads"
    if not uploads.is_dir():
        return "No uploads/ directory found."
    files = sorted(f for f in uploads.rglob("*") if f.is_file())
    if not files:
        return "uploads/ directory is empty."
    lines = [f"Files in uploads/ ({len(files)}):"]
    for f in files:
        rel = f.relative_to(uploads)
        size = f.stat().st_size
        lines.append(f"  {rel} ({_format_size(size)})")
    return "\n".join(lines)


def do_info(workspace: str, args: dict) -> str:
    """Get file metadata without full extraction."""
    file_path = _resolve_path(workspace, args)
    ext = file_path.suffix.lower()
    size = file_path.stat().st_size
    lines = [
        f"File: {file_path.name}",
        f"Size: {_format_size(size)}",
        f"Format: {ext}",
    ]
    if ext == ".pdf":
        from pypdf import PdfReader
        reader = PdfReader(str(file_path))
        lines.append(f"Pages: {len(reader.pages)}")
    elif ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(str(file_path), read_only=True)
        lines.append(f"Sheets: {', '.join(wb.sheetnames)}")
        wb.close()
    elif ext == ".csv":
        with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
            row_count = sum(1 for _ in csv.reader(f))
        lines.append(f"Rows: {row_count}")
    return "\n".join(lines)


def do_read(workspace: str, args: dict) -> str:
    """Extract text content from a file.

    Each format reader handles its own truncation at semantic boundaries
    (pages, rows, paragraphs) and includes a structural header + continuation
    hints when the output exceeds _MAX_OUTPUT_CHARS.
    """
    file_path = _resolve_path(workspace, args)
    ext = file_path.suffix.lower()

    if ext == ".pdf":
        return _read_pdf(file_path, args.get("pages"))
    elif ext == ".docx":
        return _read_docx(file_path)
    elif ext == ".xlsx":
        return _read_xlsx(file_path)
    elif ext == ".csv":
        return _read_csv(file_path)
    elif ext in _TEXT_EXTENSIONS or _is_likely_text(file_path):
        return _read_text(file_path)
    else:
        return f"Unsupported file format: {ext}. Supported: PDF, DOCX, XLSX, CSV, and plain text."


# ---------------------------------------------------------------------------
# Format readers (each handles its own smart truncation)
# ---------------------------------------------------------------------------


def _read_pdf(path: Path, pages_arg: str | None = None) -> str:
    """Extract text from a PDF file, optionally only specific pages."""
    from pypdf import PdfReader
    reader = PdfReader(str(path))
    total_pages = len(reader.pages)

    if pages_arg:
        indices = _parse_page_ranges(pages_arg, total_pages)
    else:
        indices = range(total_pages)

    # Extract pages, tracking char budget
    parts: list[str] = []
    total_chars = 0
    last_shown = 0
    for i in indices:
        text = reader.pages[i].extract_text() or ""
        if not text.strip():
            continue
        page_text = f"--- Page {i + 1} ---\n{text.strip()}"
        if total_chars + len(page_text) > _MAX_OUTPUT_CHARS and parts:
            # Budget exceeded — stop at page boundary
            last_shown = parts[-1].split("\n")[0]  # "--- Page N ---"
            break
        parts.append(page_text)
        total_chars += len(page_text)
    else:
        # Loop completed without break — all pages fit
        last_shown = None

    if not parts:
        return f"Document: {path.name} ({total_pages} pages)\nNo extractable text."

    header = f"Document: {path.name} ({total_pages} pages)"
    body = "\n\n".join(parts)

    if last_shown is not None:
        # Find the last page number shown
        last_page_shown = 0
        for p in reversed(parts):
            first_line = p.split("\n")[0]
            if first_line.startswith("--- Page "):
                last_page_shown = int(first_line.split()[2])
                break
        next_start = last_page_shown + 1
        next_end = min(last_page_shown + 10, total_pages)
        hint = (
            f"\n\nShowing pages 1-{last_page_shown} of {total_pages}. "
            f'Use pages="{next_start}-{next_end}" to read more.'
        )
        if pages_arg:
            hint = (
                f"\n\nShowing {len(parts)} of {len(list(indices)) if not isinstance(indices, range) else len(indices)} requested pages (budget reached). "
                f'Use pages="{next_start}-{next_end}" to read more.'
            )
        return f"{header}\n\n{body}{hint}"

    return f"{header}\n\n{body}"


def _read_docx(path: Path) -> str:
    """Extract text from a DOCX file."""
    from docx import Document
    doc = Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    if not paragraphs:
        return f"Document: {path.name}\nDOCX file has no text content."

    total_text = "\n\n".join(paragraphs)
    total_chars = len(total_text)
    header = f"Document: {path.name} (~{total_chars} chars)"

    if total_chars <= _MAX_OUTPUT_CHARS:
        return f"{header}\n\n{total_text}"

    # Truncate at paragraph boundary
    kept: list[str] = []
    chars = 0
    for para in paragraphs:
        if chars + len(para) > _MAX_OUTPUT_CHARS and kept:
            break
        kept.append(para)
        chars += len(para)
    shown_chars = sum(len(p) for p in kept)
    body = "\n\n".join(kept)
    hint = (
        f"\n\nShowing first {shown_chars} of {total_chars} chars. "
        f"Use exec tasks (head, grep) on the file for specific sections."
    )
    return f"{header}\n\n{body}{hint}"


def _read_xlsx(path: Path) -> str:
    """Extract text from an XLSX file (all sheets)."""
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    header = f"Workbook: {path.name} ({len(sheet_names)} sheets: {', '.join(sheet_names)})"

    parts: list[str] = []
    total_chars = 0
    truncated_sheet = None
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        rows: list[str] = []
        sheet_chars = 0
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if not any(cells):
                continue
            line = "\t".join(cells)
            row_count += 1
            if total_chars + sheet_chars + len(line) > _MAX_OUTPUT_CHARS and (parts or rows):
                # Budget exceeded mid-sheet
                truncated_sheet = (sheet_name, len(rows), row_count)
                break
            rows.append(line)
            sheet_chars += len(line) + 1  # +1 for newline

        if rows:
            parts.append(f"--- Sheet: {sheet_name} ({len(rows)} rows) ---\n" + "\n".join(rows))
            total_chars += sheet_chars

        if truncated_sheet:
            break

    wb.close()

    if not parts:
        return f"{header}\nXLSX file has no data."

    body = "\n\n".join(parts)

    if truncated_sheet:
        sname, shown, _total = truncated_sheet
        hint = (
            f"\n\nOutput truncated in sheet '{sname}' at row {shown}. "
            f'Use search(query) to find specific data across sheets.'
        )
        return f"{header}\n\n{body}{hint}"

    return f"{header}\n\n{body}"


def _read_csv(path: Path) -> str:
    """Extract text from a CSV file."""
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        all_rows = list(reader)

    if not all_rows:
        return f"Dataset: {path.name}\nCSV file is empty."

    total_rows = len(all_rows)
    # Extract column names from first row
    col_names = ", ".join(all_rows[0]) if all_rows else ""
    n_cols = len(all_rows[0]) if all_rows else 0
    header = f"Dataset: {path.name} ({total_rows} rows, {n_cols} columns)\nColumns: {col_names}"

    lines: list[str] = []
    total_chars = 0
    for row in all_rows:
        line = "\t".join(row)
        if total_chars + len(line) > _MAX_OUTPUT_CHARS and lines:
            break
        lines.append(line)
        total_chars += len(line) + 1
    else:
        # All rows fit
        body = "\n".join(lines)
        return f"{header}\n\n{body}"

    shown_rows = len(lines)
    body = "\n".join(lines)
    hint = (
        f"\n\nShowing rows 1-{shown_rows} of {total_rows}. "
        f"Use search(query) to find specific data."
    )
    return f"{header}\n\n{body}{hint}"


def _read_text(path: Path) -> str:
    """Read a plain text file."""
    text = path.read_text(encoding="utf-8", errors="replace")
    total_chars = len(text)
    header = f"Document: {path.name} (~{total_chars} chars)"

    if total_chars <= _MAX_OUTPUT_CHARS:
        return f"{header}\n\n{text}"

    # Truncate at line boundary
    truncated = text[:_MAX_OUTPUT_CHARS]
    last_newline = truncated.rfind("\n")
    if last_newline > 0:
        truncated = truncated[:last_newline]
    shown_chars = len(truncated)
    hint = (
        f"\n\nShowing first {shown_chars} of {total_chars} chars. "
        f"Use exec tasks (head, grep) on the file for specific sections."
    )
    return f"{header}\n\n{truncated}{hint}"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _resolve_path(workspace: str, args: dict) -> Path:
    """Resolve file_path arg to an absolute Path."""
    file_path = args.get("file_path")
    if not file_path:
        raise ValueError("file_path argument is required for read/info actions")
    resolved = (Path(workspace) / file_path).resolve()
    # Path traversal guard
    ws_resolved = Path(workspace).resolve()
    if not str(resolved).startswith(str(ws_resolved)):
        raise ValueError(f"Path traversal denied: {file_path}")
    if not resolved.is_file():
        raise FileNotFoundError(resolved.name)
    return resolved


def _parse_page_ranges(spec: str, total: int) -> list[int]:
    """Parse page range spec like '1-5,7,10-12' into zero-based indices."""
    indices: list[int] = []
    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-", 1)
            start_i = max(0, int(start) - 1)
            end_i = min(total, int(end))
            indices.extend(range(start_i, end_i))
        else:
            i = int(part) - 1
            if 0 <= i < total:
                indices.append(i)
    return sorted(set(indices))


def _format_size(size: int) -> str:
    """Format byte size as human-readable string."""
    if size < 1024:
        return f"{size} B"
    if size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    return f"{size / (1024 * 1024):.1f} MB"


def _is_likely_text(path: Path) -> bool:
    """Heuristic: check if the first 512 bytes look like text."""
    try:
        sample = path.read_bytes()[:512]
        # If mostly printable ASCII + whitespace, treat as text
        text_chars = sum(1 for b in sample if 32 <= b < 127 or b in (9, 10, 13))
        return len(sample) > 0 and text_chars / len(sample) > 0.85
    except Exception:
        return False


if __name__ == "__main__":
    main()
